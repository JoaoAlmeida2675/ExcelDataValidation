import openpyxl

# Function to get the column index from the column letter (A=1, B=2, etc.) - supports columns with multiple letters (e.g., AA, AB)
def get_column_index_from_letter(letter):
    letter = letter.upper()
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    column_index = 0

    for char in letter:
        char_index = alphabet.index(char) + 1
        column_index = column_index * 26 + char_index

    return column_index

# Function to generate a range including the end value
def inclusive_range(start, end):
    return range(start, end + 1)

# Open Excel and load the file
excel = openpyxl.load_workbook("your\path\to\File.xlsx")
if excel is None:
    print("Failed to open the workbook.")
else:
    # Specify the sheet name
    sheet_name = "your\sheet\name"

    # Select the sheet by name
    sheet = excel[sheet_name]

    if sheet is None:
        print("Failed to select the worksheet.")
    else:
        # Specify the columns to check (based on column letters)
        columns_to_check = ["B", "C", "D", "H", "I", "L", "M", "N", "O", "Q", "R", "T", "U", "Y", "Z", "AA", "AD", "AH", "AI", "AJ", "AK", "AM", "AN", "AO"]

        # Specify the rows to check
        rows_to_check = list(inclusive_range(6, 10))  # Check rows from 6 to 10 (including 10)

        # Specify the row with names
        row_with_names = 1  # Update with the actual row number containing column names

        # Initialize a list to store the names of empty columns
        empty_column_names = []

        # Loop through columns and rows and get the column name at the specified row
        for column_letter in columns_to_check:
            column_index = get_column_index_from_letter(column_letter)

            if column_index == 0:
                print(f"Invalid column letter: {column_letter}")
                continue

            # Get the cell at the specified row and column
            cell = sheet.cell(row=row_with_names, column=column_index)

            # Check for null values when accessing the cell
            if cell.value is None:
                column_name = "N/A"  # or any default value
            else:
                column_name = cell.value

            # Loop through rows and check for missing data
            for row in rows_to_check:
                cell_value = sheet.cell(row=row, column=column_index).value

                if cell_value is None:
                    message = f"Row: {row} - Missing Data in Column: {column_name}"
                    print(message)
                    empty_column_names.append(message)

        # Create the .txt file with messages of empty columns
        file_name = "your\save\path\file\name.txt"
        with open(file_name, "w") as file:
            file.write("\n".join(empty_column_names))

        # Close Excel
        excel.close()
